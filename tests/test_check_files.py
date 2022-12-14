import csv
import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from zipfile import ZipFile
from io import TextIOWrapper
from create_arc import create_arc, delete_arc, zip_path


arc: ZipFile
csv_file = 'username.csv'
pdf_file = 'istqb_ctfl_exam_structure_and_rules_v1.2.pdf'
xlsx_file = 'Матрица компетенций QA.xlsx'


@pytest.fixture(scope='session', autouse=True)
def prepare_data():
    delete_arc()
    create_arc()
    global arc
    arc = ZipFile(zip_path)
    yield


def test_read_csv():
    with arc.open(csv_file) as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        new_list = []
        for row in csvfile:
            if row:
                new_list.append(row)
        assert new_list == [['Username; Identifier;First name;Last name'],
                            ['booker12;9012;Rachel;Booker'],
                            ['grey07;2070;Laura;Grey'],
                            ['johnson81;4081;Craig;Johnson'],
                            ['jenkins46;9346;Mary;Jenkins'],
                            ['smith79;5079;Jamie;Smith']]


def test_read_pdf():
    with arc.open(pdf_file) as pdffile:
        reader = PdfReader(pdffile)
        page = reader.pages[1]
        text = page.extract_text()
        assert 'ISTQB - Foundation Level Exam Structure & Rules' in text


def test_read_xlsx():
    with arc.open(xlsx_file) as xfile:
        workbook = load_workbook(xfile)
        sheet = workbook.active
        assert sheet.cell(row=14, column=2).value == 'Работа с bash'
